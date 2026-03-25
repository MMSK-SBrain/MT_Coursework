#!/usr/bin/env python3
"""
Generic YAML to Excel Converter for Course Learning Outcomes
Works with any course following the standard YAML structure

Usage:
    python yaml_to_excel.py <path_to_yaml_file> [output_excel_file]

Examples:
    python scripts/yaml_to_excel.py courses/automotive-vehicles-aelzc441/course-data.yaml
    python scripts/yaml_to_excel.py courses/my-course/course-data.yaml output.xlsx
"""

import sys
import os
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def main():
    # Parse command line arguments
    if len(sys.argv) < 2:
        print("Usage: python yaml_to_excel.py <path_to_yaml_file> [output_excel_file]")
        print("\nExample:")
        print("  python scripts/yaml_to_excel.py courses/automotive-vehicles-aelzc441/course-data.yaml")
        sys.exit(1)

    yaml_file = sys.argv[1]

    # Determine output file name
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # Default: same directory as YAML, named course-template.xlsx
        yaml_dir = os.path.dirname(yaml_file)
        output_file = os.path.join(yaml_dir, "course-template.xlsx")

    # Verify YAML file exists
    if not os.path.exists(yaml_file):
        print(f"Error: YAML file not found: {yaml_file}")
        sys.exit(1)

    # Read YAML file
    print(f"Reading YAML file: {yaml_file}")
    try:
        with open(yaml_file, 'r') as f:
            data = yaml.safe_load(f)
    except Exception as e:
        print(f"Error reading YAML file: {e}")
        sys.exit(1)

    # Create workbook
    print("Creating Excel workbook...")
    wb = openpyxl.Workbook()

    # ============================================================================
    # SHEET 1: Course_Info
    # ============================================================================
    ws_info = wb.active
    ws_info.title = "Course_Info"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Add course info
    course = data['course']
    course_info = [
        ["Field", "Value"],
        ["Course Name", course.get('name', '')],
        ["Course Code", course.get('code', '')],
        ["Duration (Days)", str(course.get('duration_days', ''))],
        ["Description", course.get('description', '')],
        ["Institution", course.get('institution', '')],
        ["Credits", str(course.get('credits', ''))],
        ["Total Sessions", str(course.get('total_sessions', ''))],
        ["Session Duration", course.get('session_duration', '')],
    ]

    # Add pedagogical approach if present
    if 'pedagogical_approach' in course:
        course_info.append(["Pedagogical Approach", course['pedagogical_approach']])

    for row_idx, row_data in enumerate(course_info, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 100

    # ============================================================================
    # SHEET 2: Outcomes_And_Structure
    # ============================================================================
    ws_structure = wb.create_sheet("Outcomes_And_Structure")

    # Define column headers
    headers = [
        "Module", "Module Name", "Module Desc",
        "Lesson", "Lesson Name", "Day",
        "CO ID", "CO Code", "CO Desc", "CO Category", "CO Bloom",
        "MO ID", "MO Code", "MO Desc", "MO Category", "MO Bloom",
        "SO ID", "SO Code", "SO Desc", "SO Category", "SO Bloom",
        "Session Type", "Required"
    ]

    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_structure.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Create CO lookup dictionary
    co_lookup = {co['id']: co for co in data['course_outcomes']}

    # Add data rows
    row_num = 2

    for module in data['modules']:
        mo = module['module_outcome']
        parent_co = co_lookup[mo['parent_co']]

        for lesson in module['lessons']:
            first_so_in_lesson = True

            for so in lesson['session_outcomes']:
                # Determine what to fill based on first SO in lesson
                if first_so_in_lesson:
                    # Fill all columns for first SO
                    row_data = [
                        module['number'],
                        module['name'],
                        module['description'],
                        lesson['number'],
                        lesson['name'],
                        lesson['day'],
                        parent_co['id'],
                        parent_co['code'],
                        parent_co['description'],
                        parent_co['category'],
                        parent_co['bloom_level'],
                        mo['id'],
                        mo['code'],
                        mo['description'],
                        mo['category'],
                        mo['bloom_level'],
                        so['id'],
                        so['code'],
                        so['description'],
                        so['category'],
                        so['bloom_level'],
                        so['session_type'],
                        "Yes" if so['required'] else "No"
                    ]
                    first_so_in_lesson = False
                else:
                    # Only fill SO columns (inherit module, lesson, CO, MO)
                    row_data = [
                        "", "", "",  # Module columns (inherit)
                        "", "", "",  # Lesson columns (inherit)
                        "", "", "", "", "",  # CO columns (inherit)
                        "", "", "", "", "",  # MO columns (inherit)
                        so['id'],
                        so['code'],
                        so['description'],
                        so['category'],
                        so['bloom_level'],
                        so['session_type'],
                        "Yes" if so['required'] else "No"
                    ]

                # Write row
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_structure.cell(row=row_num, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                row_num += 1

    # Set column widths for structure sheet
    column_widths = {
        'A': 8,   # Module number
        'B': 35,  # Module Name
        'C': 60,  # Module Desc
        'D': 8,   # Lesson number
        'E': 40,  # Lesson Name
        'F': 6,   # Day
        'G': 10,  # CO ID
        'H': 18,  # CO Code
        'I': 60,  # CO Desc
        'J': 12,  # CO Category
        'K': 12,  # CO Bloom
        'L': 10,  # MO ID
        'M': 18,  # MO Code
        'N': 60,  # MO Desc
        'O': 12,  # MO Category
        'P': 12,  # MO Bloom
        'Q': 12,  # SO ID
        'R': 20,  # SO Code
        'S': 70,  # SO Desc
        'T': 12,  # SO Category
        'U': 12,  # SO Bloom
        'V': 12,  # Session Type
        'W': 10,  # Required
    }

    for col_letter, width in column_widths.items():
        ws_structure.column_dimensions[col_letter].width = width

    # Freeze header row
    ws_structure.freeze_panes = "A2"

    # Save workbook
    wb.save(output_file)

    # Print summary
    print(f"\n{'='*60}")
    print(f"✓ Excel template created: {output_file}")
    print(f"{'='*60}")
    print(f"Course: {course.get('name', 'N/A')} ({course.get('code', 'N/A')})")
    print(f"Total data rows: {row_num - 1}")
    print(f"Modules: {len(data['modules'])}")
    print(f"Sessions: {sum(len(m['lessons']) for m in data['modules'])}")

    if 'summary' in data:
        print(f"Total session outcomes: {data['summary'].get('total_session_outcomes', 'N/A')}")

    print(f"\nSheets created:")
    print(f"  1. Course_Info - Course metadata")
    print(f"  2. Outcomes_And_Structure - Complete CO→MO→SO hierarchy")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
