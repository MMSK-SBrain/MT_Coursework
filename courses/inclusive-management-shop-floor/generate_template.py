#!/usr/bin/env python3
"""
Generate Excel template for Inclusive Management course
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# ============================================================================
# SHEET 1: Course_Info
# ============================================================================
ws_info = wb.active
ws_info.title = "Course_Info"

# Header styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

# Add course info
course_info = [
    ["Field", "Value"],
    ["Course Name", "Inclusive Management: Building Respectful Shop Floor Leadership"],
    ["Course Code", "IMSF"],
    ["Duration (Days)", 1],
    ["Description", "Essential training for male managers on professional interactions, empathy, and inclusive management practices for women employees in shop floor environments"],
    ["Organization ID", ""],  # To be filled by user
]

for row_idx, row_data in enumerate(course_info, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:  # Header row
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Adjust column widths
ws_info.column_dimensions['A'].width = 20
ws_info.column_dimensions['B'].width = 80

# ============================================================================
# SHEET 2: Outcomes_And_Structure
# ============================================================================
ws_structure = wb.create_sheet("Outcomes_And_Structure")

# Define headers
headers = [
    "Module", "Module Name", "Module Desc",
    "Lesson", "Lesson Name", "Day",
    "CO ID", "CO Code", "CO Desc", "CO Category", "CO Bloom",
    "MO ID", "MO Code", "MO Desc", "MO Category", "MO Bloom",
    "SO ID", "SO Code", "SO Desc", "SO Category", "SO Bloom",
    "Session Type", "Required"
]

# Write headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws_structure.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
col_widths = [8, 25, 40, 8, 30, 6, 10, 15, 50, 12, 12, 12, 15, 50, 12, 12, 15, 18, 50, 12, 12, 15, 10]
for i, width in enumerate(col_widths, start=1):
    ws_structure.column_dimensions[get_column_letter(i)].width = width

# Course data structure
# Format: [Module, Module Name, Module Desc, Lesson, Lesson Name, Day, CO ID, CO Code, CO Desc, CO Cat, CO Bloom,
#          MO ID, MO Code, MO Desc, MO Cat, MO Bloom, SO ID, SO Code, SO Desc, SO Cat, SO Bloom, Session Type, Required]

data_rows = [
    # MODULE 1
    [1, "Foundation - Understanding the Context", "Understanding why inclusive management matters and establishing learning ground rules",
     1, "Why Inclusive Management Matters", 1,
     "co-1", "IMSF-CO-1", "Demonstrate respectful and inclusive management practices toward women employees in shop floor environments", "Attitude", "Apply",
     "mo-1-1", "IMSF-MO-1-1", "Understand the rationale for inclusive management practices", "Knowledge", "Understand",
     "so-1-1-1", "IMSF-SO-1-1-1", "Explain the business impact of inclusive vs exclusive management", "Knowledge", "Understand",
     "lecture", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "", "", "", "", "",
     "mixed", "Yes"],  # Video discussion for SO-1-1-1

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-1-1-2", "IMSF-SO-1-1-2", "Identify legal requirements for workplace equality in India", "Knowledge", "Remember",
     "lecture", "Yes"],

    # MODULE 2
    [2, "Building Self-Awareness", "Developing awareness of unconscious bias and emotional intelligence",
     1, "Understanding Unconscious Bias and Emotional Intelligence", 1,
     "", "", "", "", "",
     "mo-2-1", "IMSF-MO-2-1", "Recognize personal biases and emotional patterns", "Attitude", "Understand",
     "so-2-1-1", "IMSF-SO-2-1-1", "Complete personal bias self-assessment", "Attitude", "Apply",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-2-1-2", "IMSF-SO-2-1-2", "Recognize common unconscious biases in workplace settings", "Knowledge", "Understand",
     "lecture", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "", "", "", "", "",
     "mixed", "Yes"],  # Video for SO-2-1-2

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-2-1-3", "IMSF-SO-2-1-3", "Identify personal emotional triggers in management situations", "Attitude", "Analyze",
     "mixed", "Yes"],

    # MODULE 3
    [3, "Developing Empathy", "Understanding and practicing empathy in workplace contexts",
     1, "The Science and Practice of Empathy", 1,
     "co-2", "IMSF-CO-2", "Apply empathy and emotional intelligence in daily workplace interactions and conflict situations", "Skill", "Apply",
     "mo-3-1", "IMSF-MO-3-1", "Apply empathy in workplace interactions", "Skill", "Apply",
     "so-3-1-1", "IMSF-SO-3-1-1", "Define empathy and explain its neurological basis", "Knowledge", "Understand",
     "lecture", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-3-1-2", "IMSF-SO-3-1-2", "Recognize women's perspectives on shop floor challenges", "Attitude", "Understand",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-3-1-3", "IMSF-SO-3-1-3", "Apply empathy mapping to understand employee experiences", "Skill", "Apply",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-3-1-4", "IMSF-SO-3-1-4", "Demonstrate active empathetic listening techniques", "Skill", "Apply",
     "practical", "Yes"],

    # MODULE 4
    [4, "Practical Communication & Daily Interactions", "Building professional communication and interaction skills",
     1, "Professional Workplace Communication", 1,
     "", "", "", "", "",
     "mo-4-1", "IMSF-MO-4-1", "Conduct respectful daily workplace interactions", "Skill", "Apply",
     "so-4-1-1", "IMSF-SO-4-1-1", "Demonstrate appropriate professional greetings and daily interactions", "Skill", "Apply",
     "lecture", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-4-1-2", "IMSF-SO-4-1-2", "Identify respectful vs disrespectful workplace behaviors", "Knowledge", "Understand",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "", "", "", "", "",
     "", ""],  # Transition to MO-4-2

    ["", "", "", "", "", "", "", "", "", "", "",
     "mo-4-2", "IMSF-MO-4-2", "Provide constructive feedback professionally", "Skill", "Apply",
     "so-4-2-1", "IMSF-SO-4-2-1", "Practice professional communication in common scenarios", "Skill", "Apply",
     "practical", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-4-2-2", "IMSF-SO-4-2-2", "Deliver constructive feedback respectfully", "Skill", "Apply",
     "practical", "Yes"],

    # MODULE 5
    [5, "Health Sensitivity & Workplace Courtesies", "Understanding and accommodating health-related workplace needs",
     1, "Understanding and Responding to Health Needs", 1,
     "co-3", "IMSF-CO-3", "Implement appropriate workplace accommodations and support systems for diverse employee needs", "Skill", "Apply",
     "mo-5-1", "IMSF-MO-5-1", "Provide appropriate health-related accommodations", "Skill", "Apply",
     "so-5-1-1", "IMSF-SO-5-1-1", "Explain menstrual health needs in workplace context", "Knowledge", "Understand",
     "lecture", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-5-1-2", "IMSF-SO-5-1-2", "Respond professionally to health-related accommodation requests", "Skill", "Apply",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-5-1-3", "IMSF-SO-5-1-3", "Propose team norms for supportive environment", "Skill", "Create",
     "mixed", "Yes"],

    # MODULE 6
    [6, "Application & Integration", "Applying all learned skills to realistic workplace scenarios",
     1, "Applying Inclusive Management Frameworks", 1,
     "", "", "", "", "",
     "mo-6-1", "IMSF-MO-6-1", "Analyze and resolve complex workplace situations", "Skill", "Analyze",
     "so-6-1-1", "IMSF-SO-6-1-1", "Analyze workplace scenarios using inclusive management frameworks", "Skill", "Analyze",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-6-1-2", "IMSF-SO-6-1-2", "Formulate appropriate responses to complex workplace situations", "Skill", "Evaluate",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "", "", "", "", "",
     "quiz", "Yes"],

    # MODULE 7
    [7, "Commitment & Action Planning", "Creating personal commitments for ongoing development",
     1, "Personal Action Planning", 1,
     "", "", "", "", "",
     "mo-7-1", "IMSF-MO-7-1", "Commit to ongoing inclusive management development", "Attitude", "Apply",
     "so-7-1-1", "IMSF-SO-7-1-1", "Create personal action plan with specific behavioral commitments", "Skill", "Create",
     "mixed", "Yes"],

    ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
     "so-7-1-2", "IMSF-SO-7-1-2", "Articulate public commitment to inclusive practices", "Attitude", "Apply",
     "mixed", "Yes"],
]

# Write data rows
for row_idx, row_data in enumerate(data_rows, start=2):  # Start from row 2 (after headers)
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_structure.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Freeze top row
ws_structure.freeze_panes = "A2"

# ============================================================================
# SHEET 3: Reference (Valid Values)
# ============================================================================
ws_ref = wb.create_sheet("Reference")

reference_data = [
    ["Valid Categories", "Valid Bloom Levels", "Valid Session Types"],
    ["Knowledge", "Remember", "lecture"],
    ["Skill", "Understand", "quiz"],
    ["Attitude", "Apply", "lab"],
    ["", "Analyze", "practical"],
    ["", "Evaluate", "mixed"],
    ["", "Create", ""],
]

for row_idx, row_data in enumerate(reference_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_ref.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws_ref.column_dimensions['A'].width = 20
ws_ref.column_dimensions['B'].width = 20
ws_ref.column_dimensions['C'].width = 20

# Save workbook
output_file = "/Volumes/Dev/Course_Generator/courses/inclusive-management-shop-floor/03-template.xlsx"
wb.save(output_file)
print(f"Excel template created successfully: {output_file}")
