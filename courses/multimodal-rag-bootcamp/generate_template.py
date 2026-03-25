#!/usr/bin/env python3
"""
Generate Excel template for Multimodal RAG Bootcamp course
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# ============================================================================
# SHEET 1: Course_Info
# ============================================================================
ws_info = wb.active
ws_info.title = "Course_Info"

# Header styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

# Add course info
course_info = [
    ["Field", "Value"],
    ["Course Name", "Multimodal RAG Bootcamp: Build AI Apps That See, Read & Reason"],
    ["Course Code", "MRAG"],
    ["Duration (Days)", 1],
    ["Description", "A 4-hour hands-on bootcamp where learners build a working Multimodal RAG application from scratch, progressing from foundational concepts to a functional prototype that retrieves and reasons over text, images, and tables."],
    ["Organization ID", ""],  # To be filled by user
]

for row_idx, row_data in enumerate(course_info, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws_info.column_dimensions['A'].width = 20
ws_info.column_dimensions['B'].width = 80

# ============================================================================
# SHEET 2: Outcomes_And_Structure
# ============================================================================
ws_structure = wb.create_sheet("Outcomes_And_Structure")

headers = [
    "Module", "Module Name", "Module Desc",
    "Lesson", "Lesson Name", "Day",
    "CO ID", "CO Code", "CO Desc", "CO Category", "CO Bloom",
    "MO ID", "MO Code", "MO Desc", "MO Category", "MO Bloom",
    "SO ID", "SO Code", "SO Desc", "SO Category", "SO Bloom",
    "Session Type", "Required"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws_structure.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

col_widths = [8, 30, 45, 8, 35, 6, 10, 15, 55, 12, 12, 12, 18, 55, 12, 12, 18, 20, 55, 12, 12, 15, 10]
for i, width in enumerate(col_widths, start=1):
    ws_structure.column_dimensions[get_column_letter(i)].width = width

# ============================================================================
# COURSE DATA
# ============================================================================
# E = empty string for "inherit from above"
E = ""

data_rows = [
    # ========================================================================
    # MODULE 1: RAG Foundations (45 min)
    # ========================================================================
    [1, "RAG Foundations", "Understanding what RAG is, why multimodal matters, and the end-to-end architecture",
     1, "What is RAG & Why Multimodal?", 1,
     "co-1", "MRAG-CO-1", "Build a functional Multimodal RAG application that retrieves and reasons over text, images, and tables", "Skill", "Create",
     "mo-1-1", "MRAG-MO-1-1", "Explain the RAG architecture and articulate why multimodal retrieval is necessary for real-world AI applications", "Knowledge", "Understand",
     "so-1-1-1", "MRAG-SO-1-1-1", "Describe the limitations of standalone LLMs (hallucination, knowledge cutoff) that RAG addresses", "Knowledge", "Understand",
     "lecture", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-1-1-2", "MRAG-SO-1-1-2", "Diagram the core RAG pipeline: Ingest, Embed, Store, Retrieve, Generate", "Knowledge", "Understand",
     "lecture", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-1-1-3", "MRAG-SO-1-1-3", "Identify real-world use cases requiring multimodal retrieval (documents with images, tables, diagrams)", "Knowledge", "Understand",
     "mixed", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-1-1-4", "MRAG-SO-1-1-4", "Set up the development environment (Colab notebook, install dependencies, verify API access)", "Skill", "Apply",
     "lab", "Yes"],

    # ========================================================================
    # MODULE 2: Data Ingestion & Embedding (60 min)
    # ========================================================================
    # Lesson 2.1: Document Parsing & Chunking (30 min)
    [2, "Data Ingestion & Embedding", "Loading multimodal documents, parsing, chunking, and generating vector embeddings",
     1, "Document Parsing & Chunking", 1,
     E, E, E, E, E,
     "mo-2-1", "MRAG-MO-2-1", "Parse multimodal documents into structured chunks suitable for embedding", "Skill", "Apply",
     "so-2-1-1", "MRAG-SO-2-1-1", "Load and parse PDF documents containing text, images, and tables using Unstructured.io", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-2-1-2", "MRAG-SO-2-1-2", "Apply chunking strategies (recursive text splitting, element-aware chunking) to parsed content", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-2-1-3", "MRAG-SO-2-1-3", "Inspect and validate parsed output to confirm text, images, and tables are correctly extracted", "Skill", "Analyze",
     "lab", "Yes"],

    # Lesson 2.2: Multimodal Embeddings (30 min)
    [E, E, E,
     2, "Multimodal Embeddings", 1,
     E, E, E, E, E,
     "mo-2-2", "MRAG-MO-2-2", "Generate vector embeddings for text, image, and table content using appropriate models", "Skill", "Apply",
     "so-2-2-1", "MRAG-SO-2-2-1", "Explain how embedding models convert text and images into vector representations", "Knowledge", "Understand",
     "lecture", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-2-2-2", "MRAG-SO-2-2-2", "Generate embeddings for text chunks using OpenAI or open-source embedding models", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-2-2-3", "MRAG-SO-2-2-3", "Generate embeddings for images using multimodal embedding models (CLIP or vision model summaries)", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-2-2-4", "MRAG-SO-2-2-4", "Summarize tables into text descriptions and embed them for retrieval", "Skill", "Apply",
     "lab", "Yes"],

    # ========================================================================
    # MODULE 3: Vector Store & Retrieval (45 min)
    # ========================================================================
    # Lesson 3.1: Building the Vector Store (20 min)
    [3, "Vector Store & Retrieval", "Setting up ChromaDB, indexing multimodal embeddings, and performing similarity search",
     1, "Building the Vector Store", 1,
     E, E, E, E, E,
     "mo-3-1", "MRAG-MO-3-1", "Build and populate a ChromaDB vector store with multimodal embeddings", "Skill", "Apply",
     "so-3-1-1", "MRAG-SO-3-1-1", "Initialize a ChromaDB collection with appropriate distance metrics for multimodal data", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-3-1-2", "MRAG-SO-3-1-2", "Index text, image, and table embeddings into the vector store with metadata", "Skill", "Apply",
     "lab", "Yes"],

    # Lesson 3.2: Retrieval & Search (25 min)
    [E, E, E,
     2, "Retrieval & Search", 1,
     E, E, E, E, E,
     "mo-3-2", "MRAG-MO-3-2", "Execute similarity searches and analyze retrieval results across modalities", "Skill", "Analyze",
     "so-3-2-1", "MRAG-SO-3-2-1", "Perform similarity search queries and interpret relevance scores", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-3-2-2", "MRAG-SO-3-2-2", "Compare retrieval results across text-only vs multimodal queries", "Skill", "Analyze",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-3-2-3", "MRAG-SO-3-2-3", "Apply metadata filtering to improve retrieval precision", "Skill", "Apply",
     "lab", "Yes"],

    # ========================================================================
    # MODULE 4: RAG Pipeline Assembly (60 min)
    # ========================================================================
    # Lesson 4.1: Connecting Retrieval to Generation (30 min)
    [4, "RAG Pipeline Assembly", "Integrating retrieval with LLM generation using LangChain and building the capstone app",
     1, "Connecting Retrieval to Generation", 1,
     E, E, E, E, E,
     "mo-4-1", "MRAG-MO-4-1", "Construct a complete RAG chain using LangChain that integrates retrieval with LLM generation", "Skill", "Create",
     "so-4-1-1", "MRAG-SO-4-1-1", "Build a LangChain retrieval chain connecting ChromaDB retriever to an LLM", "Skill", "Apply",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-4-1-2", "MRAG-SO-4-1-2", "Write prompt templates that inject retrieved text, image descriptions, and table data as context", "Skill", "Create",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-4-1-3", "MRAG-SO-4-1-3", "Test the pipeline with sample questions and verify grounded answers", "Skill", "Analyze",
     "lab", "Yes"],

    # Lesson 4.2: Capstone - Build Your RAG App (30 min)
    [E, E, E,
     2, "Capstone: Build Your RAG App", 1,
     E, E, E, E, E,
     "mo-4-2", "MRAG-MO-4-2", "Design effective prompts that incorporate retrieved multimodal context for accurate generation", "Skill", "Create",
     "so-4-2-1", "MRAG-SO-4-2-1", "Assemble a complete end-to-end Multimodal RAG application using instructor-provided sample documents", "Skill", "Create",
     "practical", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-4-2-2", "MRAG-SO-4-2-2", "Query the application with diverse questions requiring text, image, and table retrieval", "Skill", "Apply",
     "practical", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-4-2-3", "MRAG-SO-4-2-3", "Present the working application and demonstrate a successful multimodal query-answer pair", "Skill", "Apply",
     "practical", "Yes"],

    # ========================================================================
    # MODULE 5: Evaluation & Next Steps (30 min)
    # ========================================================================
    [5, "Evaluation & Next Steps", "Evaluating RAG output quality, debugging common issues, and charting the path forward",
     1, "RAG Quality & Debugging", 1,
     "co-2", "MRAG-CO-2", "Evaluate RAG pipeline output quality and apply debugging strategies to improve retrieval accuracy", "Skill", "Evaluate",
     "mo-5-1", "MRAG-MO-5-1", "Evaluate RAG output quality and identify strategies for improvement", "Skill", "Evaluate",
     "so-5-1-1", "MRAG-SO-5-1-1", "Identify common RAG failure modes (irrelevant retrieval, context overflow, hallucination despite context)", "Knowledge", "Analyze",
     "lecture", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-5-1-2", "MRAG-SO-5-1-2", "Apply basic evaluation techniques (relevance scoring, answer faithfulness checks) to RAG output", "Skill", "Evaluate",
     "lab", "Yes"],

    [E, E, E, E, E, E, E, E, E, E, E, E, E, E, E, E,
     "so-5-1-3", "MRAG-SO-5-1-3", "List concrete next steps for advancing RAG skills (hybrid search, reranking, production deployment)", "Knowledge", "Remember",
     "lecture", "Yes"],
]

# Write data rows
for row_idx, row_data in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_structure.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Freeze top row
ws_structure.freeze_panes = "A2"

# ============================================================================
# SHEET 3: Reference (Valid Values)
# ============================================================================
ws_ref = wb.create_sheet("Reference")

reference_data = [
    ["Valid Categories", "Valid Bloom Levels", "Valid Session Types"],
    ["Knowledge", "Remember", "lecture"],
    ["Skill", "Understand", "quiz"],
    ["Attitude", "Apply", "lab"],
    ["", "Analyze", "practical"],
    ["", "Evaluate", "mixed"],
    ["", "Create", ""],
]

for row_idx, row_data in enumerate(reference_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_ref.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws_ref.column_dimensions['A'].width = 20
ws_ref.column_dimensions['B'].width = 20
ws_ref.column_dimensions['C'].width = 20

# Save workbook
output_file = "/Volumes/Dev/Course_Generator/courses/multimodal-rag-bootcamp/03-template.xlsx"
wb.save(output_file)
print(f"Excel template created successfully: {output_file}")
