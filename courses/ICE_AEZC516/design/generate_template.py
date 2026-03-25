#!/usr/bin/env python3
"""
Generate Excel template for AE ZG516 - Advances in IC Engines
Course import template for Reynlab CDS platform
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

# Course metadata
COURSE_NAME = "Advances in IC Engines"
COURSE_CODE = "AEZG516"
DURATION_DAYS = 15  # 15 contact sessions across semester
DESCRIPTION = "Performance and emissions, Engine design and operating parameters, Thermodynamic cycles, Gas exchange processes, Fluid motion, Turbocharging, Fuel injection systems, Combustion (SI & CI), Engine management, Exhaust after-treatment, OBD, Cooling and lubrication"
ORG_ID = "TO_BE_FILLED"  # Will need to be filled by instructor

# Learning outcomes data structure
# Each entry: (CO_ID, CO_Desc, CO_Category, CO_Bloom, [(MO_ID, MO_Desc, MO_Cat, MO_Bloom, [(SO_ID, SO_Desc, SO_Cat, SO_Bloom, Session_Type)])...]...)

learning_outcomes = [
    # CO-1: Engine Fundamentals and Performance Analysis
    ("co-1",
     "Analyze engine architecture, performance parameters, and calibration strategies to optimize torque, power, thermal efficiency, and fuel consumption for both SI and CI engines in real-world applications",
     "Skill", "Analyze",
     [
         # Module 1 - MO-1-1
         ("mo-1-1",
          "Analyze engine architecture choices, calculate fundamental performance parameters (displacement, BMEP, thermal efficiency), and differentiate between torque and power to select appropriate engine configurations for specific applications",
          "Skill", "Analyze",
          [
              ("so-1-1-1", "Analyze engine architecture choices (configuration, cylinders, bore/stroke) and their impact on performance and packaging", "Skill", "Analyze", "lecture"),
              ("so-1-1-2", "Differentiate between torque and power and calculate BMEP and thermal efficiency from performance data", "Skill", "Apply", "lecture"),
              ("so-1-1-3", "Apply thermodynamic cycle principles (Otto, Diesel) to explain compression ratio effects on efficiency", "Knowledge", "Apply", "lecture"),
              ("so-1-1-4", "Identify major engine components and explain their functions in the power generation process", "Knowledge", "Understand", "lecture"),
          ]
         ),
         # Module 1 - MO-1-2
         ("mo-1-2",
          "Explain the 4-stroke cycle, calculate and evaluate volumetric efficiency, describe engine control system architecture, and diagnose basic control system failures",
          "Skill", "Analyze",
          [
              ("so-1-2-1", "Explain the 4-stroke cycle in detail and analyze valve timing effects on engine breathing", "Knowledge", "Analyze", "lecture"),
              ("so-1-2-2", "Calculate volumetric efficiency and evaluate factors that affect air flow through the engine", "Skill", "Evaluate", "lecture"),
              ("so-1-2-3", "Describe VVT strategies and their benefits across different operating conditions", "Knowledge", "Understand", "lecture"),
              ("so-1-2-4", "Explain closed-loop control systems using sensors and actuators, and diagnose basic control failures", "Skill", "Analyze", "lecture"),
          ]
         ),
         # Module 2 - MO-2-1
         ("mo-2-1",
          "Calculate thermal efficiency, BSFC, and specific outputs from performance data, interpret torque and power curves, and analyze the effects of operating parameters",
          "Skill", "Analyze",
          [
              ("so-2-3-1", "Calculate thermal efficiency, BSFC, BMEP, and specific outputs from performance data", "Skill", "Apply", "lecture"),
              ("so-2-3-2", "Interpret torque and power curves to identify engine characteristics and suitability for applications", "Skill", "Analyze", "lecture"),
              ("so-2-3-3", "Analyze the effects of spark timing, injection timing, and compression ratio on performance and efficiency", "Skill", "Analyze", "lecture"),
              ("so-2-3-4", "Use BSFC maps to identify efficient operating regions and evaluate strategies for real-world operation", "Skill", "Evaluate", "lecture"),
          ]
         ),
         # Module 2 - MO-2-2
         ("mo-2-2",
          "Apply spark timing, fuel metering, VVT, injection timing, EGR, and boost control strategies to calibrate SI and CI engines",
          "Skill", "Apply",
          [
              ("so-2-4-1", "Apply spark timing and fuel metering strategies to calibrate SI engines for performance, emissions, and drivability", "Skill", "Apply", "lecture"),
              ("so-2-4-2", "Apply injection timing, EGR, and boost control strategies to calibrate CI engines", "Skill", "Apply", "lecture"),
              ("so-2-4-3", "Evaluate calibration trade-offs between performance, emissions, fuel economy, and durability", "Skill", "Evaluate", "lecture"),
              ("so-2-4-4", "Diagnose drivability issues and propose calibration solutions", "Skill", "Analyze", "lecture"),
          ]
         ),
     ]
    ),

    # CO-2: Air-Fuel System Design and Optimization
    ("co-2",
     "Design and evaluate air intake, fuel injection, and charge motion systems to maximize volumetric efficiency, mixture quality, and combustion performance while meeting power and emissions targets",
     "Skill", "Create",
     [
         # Module 3 - MO-3-1
         ("mo-3-1",
          "Analyze factors affecting volumetric efficiency, design and evaluate turbocharging systems including matching and boost control",
          "Skill", "Evaluate",
          [
              ("so-3-5-1", "Analyze factors affecting volumetric efficiency including valve flow, port design, and intake manifold tuning", "Skill", "Analyze", "lecture"),
              ("so-3-5-2", "Design and evaluate turbocharging systems including compressor/turbine matching and boost control strategies", "Skill", "Create", "lecture"),
              ("so-3-5-3", "Compare boosting technologies (turbo, VGT, supercharger, twincharging) and select for specific applications", "Skill", "Evaluate", "lecture"),
              ("so-3-5-4", "Calculate pressure ratios, airflow requirements, and evaluate intercooling effectiveness", "Skill", "Apply", "lecture"),
          ]
         ),
         # Module 3 - MO-3-2
         ("mo-3-2",
          "Compare fuel injection technologies, calculate mixture ratios, and analyze multiple injection strategies for combustion optimization",
          "Skill", "Analyze",
          [
              ("so-3-6-1", "Compare fuel injection technologies (PFI, GDI, common rail) and evaluate their advantages and trade-offs", "Skill", "Evaluate", "lecture"),
              ("so-3-6-2", "Calculate mixture ratios (AFR, λ) and determine appropriate mixtures for different operating modes", "Skill", "Apply", "lecture"),
              ("so-3-6-3", "Analyze multiple injection strategies in diesel engines and their effects on combustion and emissions", "Skill", "Analyze", "lecture"),
              ("so-3-6-4", "Evaluate fuel properties (octane, cetane) and their impact on engine operation", "Skill", "Evaluate", "lecture"),
          ]
         ),
         # Module 3 - MO-3-3
         ("mo-3-3",
          "Differentiate between swirl, tumble, and squish, design charge motion strategies, and analyze interaction with fuel injection",
          "Skill", "Create",
          [
              ("so-3-7-1", "Differentiate between swirl, tumble, and squish and explain how each is generated and utilized", "Knowledge", "Analyze", "lecture"),
              ("so-3-7-2", "Design charge motion strategies for SI engines (tumble-based) and CI engines (swirl-based)", "Skill", "Create", "lecture"),
              ("so-3-7-3", "Analyze the interaction between fuel injection, charge motion, and piston bowl design for optimal mixing", "Skill", "Analyze", "lecture"),
              ("so-3-7-4", "Identify unwanted flows (crevices, blow-by) and evaluate their impact on emissions and efficiency", "Skill", "Evaluate", "lecture"),
          ]
         ),
     ]
    ),

    # CO-3: Combustion Process Analysis and Control
    ("co-3",
     "Analyze combustion processes in SI and CI engines, diagnose combustion-related issues using thermodynamic principles and diagnostic data, and apply advanced combustion strategies",
     "Skill", "Analyze",
     [
         # Module 4 - MO-4-1
         ("mo-4-1",
          "Explain the SI combustion process, analyze engine knock including mitigation strategies, and evaluate advanced SI combustion strategies",
          "Skill", "Analyze",
          [
              ("so-4-8-1", "Explain the SI combustion process including flame development, propagation, and the factors affecting flame speed", "Knowledge", "Understand", "lecture"),
              ("so-4-8-2", "Analyze engine knock, identify factors that promote it, and apply mitigation strategies", "Skill", "Analyze", "lecture"),
              ("so-4-8-3", "Evaluate the relationship between octane rating, compression ratio, and knock tendency", "Skill", "Evaluate", "lecture"),
              ("so-4-8-4", "Assess advanced SI combustion strategies (lean burn, stratified charge, EGR) and their trade-offs", "Skill", "Evaluate", "lecture"),
          ]
         ),
         # Module 4 - MO-4-2
         ("mo-4-2",
          "Explain the four phases of diesel combustion, analyze fuel spray behavior, and evaluate the NOx-PM trade-off",
          "Skill", "Analyze",
          [
              ("so-4-9-1", "Explain the four phases of diesel combustion (ignition delay, premixed, mixing-controlled, late combustion)", "Knowledge", "Understand", "lecture"),
              ("so-4-9-2", "Analyze fuel spray behavior including atomization, penetration, and air entrainment", "Skill", "Analyze", "lecture"),
              ("so-4-9-3", "Evaluate the NOx-PM trade-off and apply in-cylinder strategies (EGR, injection pressure, timing) to optimize emissions", "Skill", "Evaluate", "lecture"),
              ("so-4-9-4", "Compare direct injection and indirect injection diesel systems", "Skill", "Analyze", "lecture"),
              ("so-4-9-5", "Assess the effects of fuel cetane rating, injection pressure, and multiple injections on diesel combustion", "Skill", "Evaluate", "lecture"),
          ]
         ),
         # Module 4 - MO-4-3
         ("mo-4-3",
          "Compare thermodynamic cycles, analyze Atkinson/Miller cycle principles, and evaluate advanced combustion modes",
          "Skill", "Evaluate",
          [
              ("so-4-10-1", "Compare ideal thermodynamic cycles (Otto, Diesel, Dual) and explain why real engines differ from ideal", "Knowledge", "Analyze", "lecture"),
              ("so-4-10-2", "Analyze Atkinson/Miller cycle principles and evaluate their implementation using variable valve timing", "Skill", "Analyze", "lecture"),
              ("so-4-10-3", "Evaluate advanced combustion modes (HCCI, RCCI, PCCI) including their benefits and challenges", "Skill", "Evaluate", "lecture"),
              ("so-4-10-4", "Assess variable compression ratio technology and its potential for efficiency and performance optimization", "Skill", "Evaluate", "lecture"),
          ]
         ),
     ]
    ),

    # CO-4: Emission Formation and After-Treatment Systems
    ("co-4",
     "Evaluate emission formation mechanisms and design integrated emission control strategies combining in-cylinder measures and after-treatment systems to achieve regulatory compliance",
     "Skill", "Create",
     [
         # Module 5 - MO-5-1
         ("mo-5-1",
          "Explain the formation mechanisms for NOx, CO, HC, and PM, and evaluate the NOx-PM trade-off to propose reduction strategies",
          "Skill", "Evaluate",
          [
              ("so-5-11-1", "Explain the formation mechanisms for NOx (thermal Zeldovich), CO (incomplete oxidation), HC (crevices, quench), and PM (soot formation/oxidation)", "Knowledge", "Understand", "lecture"),
              ("so-5-11-2", "Analyze the factors affecting each pollutant (temperature, mixture ratio, mixing, oxygen availability)", "Skill", "Analyze", "lecture"),
              ("so-5-11-3", "Evaluate the NOx-PM trade-off in diesel engines and explain why in-cylinder optimization alone is insufficient", "Skill", "Evaluate", "lecture"),
              ("so-5-11-4", "Identify emission sources for different engine operating conditions and propose in-cylinder reduction strategies", "Skill", "Create", "lecture"),
          ]
         ),
         # Module 5 - MO-5-2
         ("mo-5-2",
          "Explain Three-Way Catalyst operation, design integrated diesel after-treatment systems, and evaluate SCR systems",
          "Skill", "Create",
          [
              ("so-5-12-1", "Explain the operation of Three-Way Catalysts (TWC) and the importance of stoichiometric operation", "Knowledge", "Understand", "lecture"),
              ("so-5-12-2", "Design integrated after-treatment systems for diesel engines (DOC, DPF, SCR)", "Skill", "Create", "lecture"),
              ("so-5-12-3", "Analyze DPF regeneration strategies (passive, active) and evaluate soot/ash accumulation issues", "Skill", "Analyze", "lecture"),
              ("so-5-12-4", "Evaluate SCR systems including urea decomposition, ammonia storage, and NOx reduction efficiency", "Skill", "Evaluate", "lecture"),
              ("so-5-12-5", "Compare after-treatment technologies (TWC, LNT, SCR) for different applications", "Skill", "Evaluate", "lecture"),
          ]
         ),
         # Module 5 - MO-5-3
         ("mo-5-3",
          "Compare global emission regulations, explain OBD-II architecture, and diagnose emission system faults using diagnostic data",
          "Skill", "Analyze",
          [
              ("so-5-13-1", "Compare global emission regulations (Euro 6/7, EPA Tier 3, BS-VI, China 6) and explain Real Driving Emissions (RDE)", "Knowledge", "Analyze", "lecture"),
              ("so-5-13-2", "Explain the purpose and architecture of OBD-II systems including monitored components and malfunction thresholds", "Knowledge", "Understand", "lecture"),
              ("so-5-13-3", "Diagnose emission system faults using DTCs, freeze frame data, and live sensor data", "Skill", "Analyze", "lecture"),
              ("so-5-13-4", "Evaluate OBD-II readiness monitors and their role in emission compliance (I/M programs)", "Skill", "Evaluate", "lecture"),
              ("so-5-13-5", "Apply diagnostic workflow to troubleshoot emission-related faults and avoid common misdiagnoses", "Skill", "Apply", "lecture"),
          ]
         ),
     ]
    ),

    # CO-5: Advanced Technologies and Future Powertrain Systems
    ("co-5",
     "Assess the technical and economic feasibility of advanced engine technologies, alternate fuels, efficiency improvements, and hybrid powertrains for future mobility and industrial applications",
     "Skill", "Evaluate",
     [
         # Module 6 - MO-6-1
         ("mo-6-1",
          "Evaluate properties and combustion characteristics of alternate fuels, assess engine modifications, and analyze IC engines in future mobility",
          "Skill", "Evaluate",
          [
              ("so-6-14-1", "Evaluate properties and combustion characteristics of alternate fuels (CNG, hydrogen, ethanol, biodiesel)", "Skill", "Evaluate", "lecture"),
              ("so-6-14-2", "Assess engine modifications required for alternate fuel operation and analyze performance/emissions trade-offs", "Skill", "Evaluate", "lecture"),
              ("so-6-14-3", "Compare lifecycle carbon emissions and sustainability of different fuel pathways", "Skill", "Analyze", "lecture"),
              ("so-6-14-4", "Analyze the role of IC engines in future mobility scenarios including hybridization and synthetic fuels", "Skill", "Analyze", "lecture"),
          ]
         ),
         # Module 6 - MO-6-2
         ("mo-6-2",
          "Analyze heat transfer mechanisms, evaluate friction reduction strategies, assess waste heat recovery, and integrate efficiency technologies",
          "Skill", "Evaluate",
          [
              ("so-6-15-1", "Analyze heat transfer mechanisms in engines and evaluate strategies to reduce heat loss", "Skill", "Analyze", "lecture"),
              ("so-6-15-2", "Evaluate friction sources and apply reduction strategies including low-viscosity oils and surface coatings", "Skill", "Evaluate", "lecture"),
              ("so-6-15-3", "Assess waste heat recovery technologies (turbocompound, ORC, TEG) and their efficiency/cost trade-offs", "Skill", "Evaluate", "lecture"),
              ("so-6-15-4", "Integrate multiple technologies (downsizing, hybridization, thermal management) for system-level efficiency", "Skill", "Create", "lecture"),
          ]
         ),
     ]
    ),
]

# Module structure
modules = [
    (1, "Engine Foundations - Build Your Engine", "Gamified introduction to engine architecture, power fundamentals, breathing, and control systems", [1, 2]),
    (2, "Performance & Calibration Excellence", "Performance analysis, calibration strategies, and optimization for SI and CI engines", [3, 4]),
    (3, "Air & Fuel Systems - The Foundation of Good Combustion", "Air intake, boosting, fuel injection, mixture formation, and charge motion", [5, 6, 7]),
    (4, "Combustion Processes - Where the Magic Happens", "SI combustion, CI combustion, advanced cycles and combustion modes", [8, 9, 10]),
    (5, "Emissions - The Constraint That Defines Modern Engine Design", "Emission formation, after-treatment systems, regulations, and diagnostics", [11, 12, 13]),
    (6, "Advanced Technologies & Future Directions", "Alternate fuels, energy transition, thermal management, and efficiency optimization", [14, 15]),
]

# Lesson names for each session
lesson_names = {
    1: "Build Your Engine - Architecture, Power & Thermodynamics",
    2: "Build Your Engine - Breathing & Control Systems",
    3: "Performance Analysis & Operating Characteristics",
    4: "Engine Calibration Strategies",
    5: "Air Intake, Exhaust & Boosting Systems",
    6: "Fuel Injection Systems & Mixture Formation",
    7: "Charge Motion & In-Cylinder Flow Dynamics",
    8: "Combustion in SI Engines",
    9: "Combustion in CI Engines",
    10: "Advanced Engine Cycles & Combustion Modes",
    11: "Emission Formation Mechanisms",
    12: "Exhaust Gas After-Treatment Systems",
    13: "Regulations, Compliance, OBD & Diagnostics",
    14: "Alternate Fuels & Energy Transition",
    15: "Thermal Management, Friction Reduction & System Integration",
}

def generate_excel_template():
    """Generate the complete Excel template"""

    # Create workbook
    wb = Workbook()

    # Sheet 1: Course_Info
    ws_info = wb.active
    ws_info.title = "Course_Info"

    course_info_data = [
        ["Field", "Value"],
        ["Course Name", COURSE_NAME],
        ["Course Code", COURSE_CODE],
        ["Duration (Days)", DURATION_DAYS],
        ["Description", DESCRIPTION],
        ["Organization ID", ORG_ID],
    ]

    for row in course_info_data:
        ws_info.append(row)

    # Format Course_Info sheet
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws_info[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 100

    # Sheet 2: Outcomes_And_Structure
    ws_structure = wb.create_sheet("Outcomes_And_Structure")

    # Define column headers
    headers = [
        "Module", "Module Name", "Module Desc",
        "Lesson", "Lesson Name", "Day",
        "CO ID", "CO Code", "CO Desc", "CO Category", "CO Bloom",
        "MO ID", "MO Code", "MO Desc", "MO Category", "MO Bloom",
        "SO ID", "SO Code", "SO Desc", "SO Category", "SO Bloom",
        "Session Type", "Required"
    ]

    ws_structure.append(headers)

    # Format header row
    for cell in ws_structure[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Build rows using inheritance pattern
    rows = []

    # Track current context for inheritance
    current_module = None
    current_module_name = None
    current_module_desc = None
    current_lesson = None
    current_lesson_name = None
    current_day = None
    current_co = None
    current_co_data = None
    current_mo = None
    current_mo_data = None
    current_so = None
    current_so_data = None

    # Map sessions to modules
    session_to_module = {}
    for mod_num, mod_name, mod_desc, sessions in modules:
        for session in sessions:
            session_to_module[session] = (mod_num, mod_name, mod_desc)

    # Iterate through learning outcomes
    for co_id, co_desc, co_cat, co_bloom, mos in learning_outcomes:
        co_code = f"{COURSE_CODE}-{co_id.upper()}"

        for mo_id, mo_desc, mo_cat, mo_bloom, sos in mos:
            mo_code = f"{COURSE_CODE}-{mo_id.upper()}"

            for so_id, so_desc, so_cat, so_bloom, session_type in sos:
                so_code = f"{COURSE_CODE}-{so_id.upper()}"

                # Parse SO ID to get module and lesson numbers
                # Format: so-{module}-{lesson}-{seq}
                parts = so_id.split('-')
                module_num = int(parts[1])
                lesson_num = int(parts[2])

                # Get lesson day (sequential)
                day_num = lesson_num + sum(len(sessions) for mod_n, _, _, sessions in modules if mod_n < module_num)

                # Get module info
                module_info = None
                for mod_num, mod_name, mod_desc, _ in modules:
                    if mod_num == module_num:
                        module_info = (mod_num, mod_name, mod_desc)
                        break

                # Build row with inheritance
                row = []

                # Module columns
                if module_info != (current_module, current_module_name, current_module_desc):
                    row.extend([module_info[0], module_info[1], module_info[2]])
                    current_module, current_module_name, current_module_desc = module_info
                else:
                    row.extend(["", "", ""])  # Inherit from above

                # Lesson columns
                lesson_key = None
                for day in lesson_names:
                    if day == day_num:
                        lesson_key = day
                        break

                lesson_name_value = lesson_names.get(lesson_key, f"Session {day_num}")

                if (lesson_num, lesson_name_value, day_num) != (current_lesson, current_lesson_name, current_day):
                    row.extend([lesson_num, lesson_name_value, day_num])
                    current_lesson, current_lesson_name, current_day = lesson_num, lesson_name_value, day_num
                else:
                    row.extend(["", "", ""])  # Inherit from above

                # CO columns
                if (co_id, co_code, co_desc, co_cat, co_bloom) != current_co_data:
                    row.extend([co_id, co_code, co_desc, co_cat, co_bloom])
                    current_co_data = (co_id, co_code, co_desc, co_cat, co_bloom)
                else:
                    row.extend(["", "", "", "", ""])  # Inherit from above

                # MO columns
                if (mo_id, mo_code, mo_desc, mo_cat, mo_bloom) != current_mo_data:
                    row.extend([mo_id, mo_code, mo_desc, mo_cat, mo_bloom])
                    current_mo_data = (mo_id, mo_code, mo_desc, mo_cat, mo_bloom)
                else:
                    row.extend(["", "", "", "", ""])  # Inherit from above

                # SO columns (always new)
                row.extend([so_id, so_code, so_desc, so_cat, so_bloom])
                current_so_data = (so_id, so_code, so_desc, so_cat, so_bloom)

                # Session columns
                row.extend([session_type, "Yes"])

                rows.append(row)

    # Add all rows to worksheet
    for row in rows:
        ws_structure.append(row)

    # Set column widths
    column_widths = {
        'A': 8, 'B': 40, 'C': 50,  # Module columns
        'D': 8, 'E': 50, 'F': 6,   # Lesson columns
        'G': 10, 'H': 18, 'I': 80, 'J': 12, 'K': 12,  # CO columns
        'L': 12, 'M': 20, 'N': 80, 'O': 12, 'P': 12,  # MO columns
        'Q': 14, 'R': 22, 'S': 100, 'T': 12, 'U': 12,  # SO columns
        'V': 15, 'W': 10  # Session columns
    }

    for col, width in column_widths.items():
        ws_structure.column_dimensions[col].width = width

    # Add Reference sheet with valid values
    ws_ref = wb.create_sheet("Reference")
    ws_ref.append(["Valid Categories"])
    ws_ref.append(["Knowledge"])
    ws_ref.append(["Skill"])
    ws_ref.append(["Attitude"])
    ws_ref.append([])
    ws_ref.append(["Valid Bloom Levels"])
    ws_ref.append(["Remember"])
    ws_ref.append(["Understand"])
    ws_ref.append(["Apply"])
    ws_ref.append(["Analyze"])
    ws_ref.append(["Evaluate"])
    ws_ref.append(["Create"])
    ws_ref.append([])
    ws_ref.append(["Valid Session Types"])
    ws_ref.append(["lecture"])
    ws_ref.append(["quiz"])
    ws_ref.append(["lab"])
    ws_ref.append(["practical"])
    ws_ref.append(["mixed"])

    # Save workbook
    output_file = "/Volumes/Dev/Course_Generator/courses/ICE_AEZC516/AEZG516_Course_Template.xlsx"
    wb.save(output_file)
    print(f"✓ Excel template generated: {output_file}")
    print(f"  - Course: {COURSE_NAME} ({COURSE_CODE})")
    print(f"  - Duration: {DURATION_DAYS} sessions")
    print(f"  - Course Outcomes: {len(learning_outcomes)}")
    print(f"  - Module Outcomes: {sum(len(mos) for _, _, _, _, mos in learning_outcomes)}")
    print(f"  - Session Outcomes: {sum(len(sos) for _, _, _, _, mos in learning_outcomes for _, _, _, _, sos in mos)}")
    print(f"  - Total rows: {len(rows)}")

    return output_file

if __name__ == "__main__":
    try:
        output_file = generate_excel_template()
        print("\n✓ SUCCESS: Template ready for import!")
        print(f"\nNext steps:")
        print(f"1. Open the Excel file: {output_file}")
        print(f"2. Fill in the Organization ID in the Course_Info sheet")
        print(f"3. Review and adjust any descriptions as needed")
        print(f"4. Import using: --dry-run first, then actual import")
    except Exception as e:
        print(f"✗ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
